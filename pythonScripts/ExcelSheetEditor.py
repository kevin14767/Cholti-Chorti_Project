from openpyxl import load_workbook
#workbook we are working with 
workbook = load_workbook(filename= "Cholti_WordList.xlsx")
workbook.sheetnames
sheet = workbook.active

column_index=3;
row_index=0;

for value in sheet.iter_rows(min_col = column_index,max_col =column_index, values_only =True):
    row_index +=1;
    if value is not None:
         string = str(value)
         for i in range(len(string)-1):
            first = string[i];
            second = string[i+1];
            if(first == 'c'):
                if(second == 'i' or second == 'e'):
                    transcription = string.replace('c','s')
                    cell = sheet.cell(row= row_index, column=4)
                    cell_string = str(cell)
                    cell_string = cell_string.split(sep='.')
                    cell_string = cell_string[1].split(sep = '>');
                    print(cell_string[0])
                    sheet[cell_string[0]] = transcription
                if(second == 'a' or second == 'o' or second == 'u'):
                    transcription = string.replace('c','k')
                    cell = sheet.cell(row= row_index, column=4)
                    cell_string = str(cell)
                    cell_string = cell_string.split(sep='.')
                    cell_string = cell_string[1].split(sep = '>');
                    print(cell_string[0])
                    sheet[cell_string[0]] = transcription



workbook.save(filename="cholti_append.xlsx");
                   
                    
